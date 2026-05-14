"""
Day 17 — Export Module
Formats: CSV, Excel, JSON, Google Sheets link, PDF
Filter panel: date, category, AI score, status
Done-When: All 5 formats download correctly with filter applied
"""

import os
import json
import logging
import io
from datetime import datetime, timedelta
from typing import List, Dict, Optional
import pandas as pd
import streamlit as st

logger = logging.getLogger(__name__)


# ==========================================
# FILTER PANEL
# ==========================================
def render_export_filters(df: pd.DataFrame) -> pd.DataFrame:
    """
    Render export filter panel.
    Filters: date range, category, validation status, AI score threshold
    Returns filtered dataframe.
    """
    st.markdown("#### 🔧 Export Filters")

    col1, col2, col3, col4 = st.columns(4)

    with col1:
        # Date range filter
        date_range = st.selectbox(
            "Date Range",
            ["All Time", "Today", "Last 7 Days", "Last 30 Days", "This Month"],
            key="export_date_range"
        )

    with col2:
        # Category filter
        categories = ["All Categories"]
        if "category" in df.columns:
            categories += sorted(df["category"].dropna().unique().tolist())
        selected_category = st.selectbox(
            "Category",
            categories,
            key="export_category"
        )

    with col3:
        # Validation status filter
        status_options = ["All Status", "Valid", "Invalid", "Pending"]
        selected_status = st.selectbox(
            "Validation Status",
            status_options,
            key="export_status"
        )

    with col4:
        # AI score threshold
        min_score = st.slider(
            "Min AI Score",
            min_value=0,
            max_value=100,
            value=0,
            step=10,
            key="export_min_score",
            help="Export only leads with AI score above this value"
        )

    # Apply filters
    filtered = df.copy()

    # Date filter
    today = datetime.now()
    date_col = "scraped_date" if "scraped_date" in df.columns else None
    if date_col and date_range != "All Time":
        try:
            filtered[date_col] = filtered[date_col].astype(str)
            if date_range == "Today":
                date_str = today.strftime("%Y-%m-%d")
                filtered = filtered[filtered[date_col].str.startswith(date_str)]
            elif date_range == "Last 7 Days":
                cutoff = (today - timedelta(days=7)).strftime("%Y-%m-%d")
                filtered = filtered[filtered[date_col] >= cutoff]
            elif date_range == "Last 30 Days":
                cutoff = (today - timedelta(days=30)).strftime("%Y-%m-%d")
                filtered = filtered[filtered[date_col] >= cutoff]
            elif date_range == "This Month":
                month_str = today.strftime("%Y-%m")
                filtered = filtered[filtered[date_col].str.startswith(month_str)]
        except Exception as e:
            logger.debug(f"Date filter error: {e}")

    # Category filter
    if selected_category != "All Categories" and "category" in filtered.columns:
        filtered = filtered[filtered["category"] == selected_category]

    # Status filter
    if selected_status != "All Status" and "validation_status" in filtered.columns:
        filtered = filtered[filtered["validation_status"] == selected_status]

    # AI score filter
    if min_score > 0 and "ai_score" in filtered.columns:
        try:
            filtered["ai_score"] = pd.to_numeric(filtered["ai_score"], errors="coerce").fillna(0)
            filtered = filtered[filtered["ai_score"] >= min_score]
        except Exception as e:
            logger.debug(f"Score filter error: {e}")

    st.info(f"📊 **{len(filtered)}** leads match your filters (from {len(df)} total)")
    return filtered


# ==========================================
# EXPORT FORMAT 1 — CSV
# ==========================================
def export_csv(df: pd.DataFrame) -> bytes:
    """Export leads as CSV file."""
    return df.to_csv(index=False).encode("utf-8")


# ==========================================
# EXPORT FORMAT 2 — Excel (formatted)
# ==========================================
def export_excel(df: pd.DataFrame) -> bytes:
    """
    Export leads as formatted Excel file using openpyxl.
    Includes headers, filters, column widths.
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import (
            Font, PatternFill, Alignment,
            Border, Side
        )
        from openpyxl.utils import get_column_letter

        wb = Workbook()
        ws = wb.active
        ws.title = "LeadPulse Leads"

        # Header style
        header_fill = PatternFill(
            start_color="2563EB",
            end_color="2563EB",
            fill_type="solid"
        )
        header_font = Font(
            color="FFFFFF",
            bold=True,
            size=11
        )
        header_alignment = Alignment(
            horizontal="center",
            vertical="center"
        )
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin")
        )

        # Write headers
        headers = list(df.columns)
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header.upper().replace("_", " "))
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
            cell.border = thin_border

        # Write data rows
        valid_fill = PatternFill(start_color="DCFCE7", end_color="DCFCE7", fill_type="solid")
        invalid_fill = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")
        pending_fill = PatternFill(start_color="FEF9C3", end_color="FEF9C3", fill_type="solid")

        for row_num, row in enumerate(df.itertuples(index=False), 2):
            for col_num, value in enumerate(row, 1):
                cell = ws.cell(row=row_num, column=col_num, value=str(value) if value is not None else "")
                cell.border = thin_border
                cell.alignment = Alignment(vertical="center", wrap_text=False)

                # Color rows by validation status
                if "validation_status" in headers:
                    status_idx = headers.index("validation_status")
                    status_val = str(row[status_idx]) if row[status_idx] else ""
                    if status_val == "Valid":
                        cell.fill = valid_fill
                    elif status_val == "Invalid":
                        cell.fill = invalid_fill
                    elif status_val == "Pending":
                        cell.fill = pending_fill

        # Auto-width columns
        for col_num, header in enumerate(headers, 1):
            max_len = max(
                len(str(header)),
                df[header].astype(str).str.len().max() if not df.empty else 0
            )
            ws.column_dimensions[get_column_letter(col_num)].width = min(max_len + 4, 40)

        # Freeze header row
        ws.freeze_panes = "A2"

        # Add auto filter
        ws.auto_filter.ref = ws.dimensions

        # Save to bytes
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output.getvalue()

    except ImportError:
        logger.error("openpyxl not installed")
        return df.to_csv(index=False).encode("utf-8")
    except Exception as e:
        logger.error(f"Excel export error: {e}")
        return df.to_csv(index=False).encode("utf-8")


# ==========================================
# EXPORT FORMAT 3 — JSON
# ==========================================
def export_json(df: pd.DataFrame) -> bytes:
    """Export leads as JSON file."""
    return df.to_json(orient="records", indent=2).encode("utf-8")


# ==========================================
# EXPORT FORMAT 4 — Google Sheets Link
# ==========================================
def get_google_sheets_link() -> str:
    """Get Google Sheets sharing link from environment."""
    try:
        # Try to get from google_sheets module directly
        import google_sheets
        # Check if module has sheet URL stored
        if hasattr(google_sheets, "SHEET_URL"):
            return google_sheets.SHEET_URL
        if hasattr(google_sheets, "get_sheet_url"):
            return google_sheets.get_sheet_url()

        # Get spreadsheet ID from credentials
        import os
        import json
        creds_json = os.environ.get("GOOGLE_SHEETS_CREDENTIALS", "")
        sheet_name = os.environ.get("SHEET_NAME", "LeadPulse_Data")

        if creds_json:
            try:
                import gspread
                from google.oauth2.service_account import Credentials
                creds_dict = json.loads(creds_json)
                scopes = [
                    "https://spreadsheets.google.com/feeds",
                    "https://www.googleapis.com/auth/drive"
                ]
                creds = Credentials.from_service_account_info(
                    creds_dict, scopes=scopes
                )
                client = gspread.authorize(creds)
                sheet = client.open(sheet_name)
                return f"https://docs.google.com/spreadsheets/d/{sheet.id}"
            except Exception as e:
                logger.debug(f"Could not get sheet ID: {e}")

        return ""
    except Exception as e:
        logger.error(f"Sheets link error: {e}")
        return ""


# ==========================================
# EXPORT FORMAT 5 — PDF Report
# ==========================================
def export_pdf(df: pd.DataFrame) -> bytes:
    """
    Export leads as formatted PDF report using reportlab.
    """
    try:
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib import colors
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import inch
        from reportlab.platypus import (
            SimpleDocTemplate, Table, TableStyle,
            Paragraph, Spacer
        )

        output = io.BytesIO()
        doc = SimpleDocTemplate(
            output,
            pagesize=landscape(A4),
            rightMargin=0.5*inch,
            leftMargin=0.5*inch,
            topMargin=0.5*inch,
            bottomMargin=0.5*inch
        )

        styles = getSampleStyleSheet()
        elements = []

        # Title
        title_style = ParagraphStyle(
            "Title",
            parent=styles["Heading1"],
            fontSize=18,
            textColor=colors.HexColor("#2563EB"),
            spaceAfter=12
        )
        elements.append(Paragraph("🚀 LeadPulse Pro — Lead Report", title_style))

        # Subtitle
        subtitle_style = ParagraphStyle(
            "Subtitle",
            parent=styles["Normal"],
            fontSize=10,
            textColor=colors.grey,
            spaceAfter=20
        )
        elements.append(Paragraph(
            f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')} | Total Leads: {len(df)}",
            subtitle_style
        ))
        elements.append(Spacer(1, 0.2*inch))

        # Summary stats
        if "validation_status" in df.columns:
            valid = len(df[df["validation_status"] == "Valid"])
            invalid = len(df[df["validation_status"] == "Invalid"])
            pending = len(df[df["validation_status"] == "Pending"])
            summary_data = [
                ["Total Leads", "Valid", "Invalid", "Pending", "Quality %"],
                [
                    str(len(df)),
                    str(valid),
                    str(invalid),
                    str(pending),
                    f"{int(valid/len(df)*100)}%" if len(df) > 0 else "0%"
                ]
            ]
            summary_table = Table(summary_data, colWidths=[2*inch]*5)
            summary_table.setStyle(TableStyle([
                ("BACKGROUND", (0,0), (-1,0), colors.HexColor("#2563EB")),
                ("TEXTCOLOR", (0,0), (-1,0), colors.white),
                ("FONTNAME", (0,0), (-1,0), "Helvetica-Bold"),
                ("FONTSIZE", (0,0), (-1,-1), 10),
                ("ALIGN", (0,0), (-1,-1), "CENTER"),
                ("GRID", (0,0), (-1,-1), 0.5, colors.grey),
                ("ROWBACKGROUNDS", (0,1), (-1,-1), [colors.white, colors.HexColor("#F8FAFC")]),
            ]))
            elements.append(summary_table)
            elements.append(Spacer(1, 0.3*inch))

        # Main data table
        display_cols = ["name", "phone", "email", "category", "rating", "validation_status"]
        available_cols = [c for c in display_cols if c in df.columns]

        if available_cols:
            table_data = [
                [col.upper().replace("_", " ") for col in available_cols]
            ]
            for _, row in df[available_cols].head(100).iterrows():
                table_data.append([
                    str(row[col])[:30] if row[col] is not None else ""
                    for col in available_cols
                ])

            col_width = (landscape(A4)[0] - inch) / len(available_cols)
            data_table = Table(table_data, colWidths=[col_width]*len(available_cols))
            data_table.setStyle(TableStyle([
                ("BACKGROUND", (0,0), (-1,0), colors.HexColor("#0F172A")),
                ("TEXTCOLOR", (0,0), (-1,0), colors.white),
                ("FONTNAME", (0,0), (-1,0), "Helvetica-Bold"),
                ("FONTSIZE", (0,0), (-1,-1), 8),
                ("ALIGN", (0,0), (-1,-1), "CENTER"),
                ("GRID", (0,0), (-1,-1), 0.25, colors.grey),
                ("ROWBACKGROUNDS", (0,1), (-1,-1), [colors.white, colors.HexColor("#F1F5F9")]),
                ("TOPPADDING", (0,0), (-1,-1), 4),
                ("BOTTOMPADDING", (0,0), (-1,-1), 4),
            ]))
            elements.append(data_table)

        doc.build(elements)
        output.seek(0)
        return output.getvalue()

    except ImportError:
        logger.error("reportlab not installed")
        return b""
    except Exception as e:
        logger.error(f"PDF export error: {e}")
        return b""


# ==========================================
# MAIN EXPORT UI
# ==========================================
def render_export_ui(df: pd.DataFrame, title: str = "Export Leads") -> None:
    """
    Render complete export UI with filters and all 5 format buttons.
    """
    st.markdown(f"### 📤 {title}")

    if df.empty:
        st.info("No leads to export. Generate leads first.")
        return

    # Apply filters
    filtered_df = render_export_filters(df)

    if filtered_df.empty:
        st.warning("No leads match your filter criteria.")
        return

    st.markdown("---")
    st.markdown("#### 📥 Download Formats")

    # Export buttons row
    col1, col2, col3, col4, col5 = st.columns(5)

    # Format 1 — CSV
    with col1:
        csv_data = export_csv(filtered_df)
        st.download_button(
            label="📄 CSV",
            data=csv_data,
            file_name=f"leadpulse_export_{datetime.now().strftime('%Y%m%d_%H%M')}.csv",
            mime="text/csv",
            use_container_width=True,
            help="Universal CRM import format"
        )

    # Format 2 — Excel
    with col2:
        excel_data = export_excel(filtered_df)
        st.download_button(
            label="📊 Excel",
            data=excel_data,
            file_name=f"leadpulse_export_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
            help="Formatted spreadsheet with headers and filters"
        )

    # Format 3 — JSON
    with col3:
        json_data = export_json(filtered_df)
        st.download_button(
            label="🔧 JSON",
            data=json_data,
            file_name=f"leadpulse_export_{datetime.now().strftime('%Y%m%d_%H%M')}.json",
            mime="application/json",
            use_container_width=True,
            help="Developer integration format"
        )

    # Format 4 — Google Sheets
    with col4:
        sheets_url = get_google_sheets_link()
        if sheets_url:
            st.link_button(
                "☁️ Google Sheet",
                sheets_url,
                use_container_width=True
            )
        else:
            st.button(
                "☁️ Google Sheet",
                disabled=True,
                use_container_width=True,
                help="Connect Google Sheets in System Settings"
            )

    # Format 5 — PDF
    with col5:
        pdf_data = export_pdf(filtered_df)
        if pdf_data:
            st.download_button(
                label="📑 PDF Report",
                data=pdf_data,
                file_name=f"leadpulse_report_{datetime.now().strftime('%Y%m%d_%H%M')}.pdf",
                mime="application/pdf",
                use_container_width=True,
                help="Formatted report for clients"
            )
        else:
            st.button(
                "📑 PDF Report",
                disabled=True,
                use_container_width=True,
                help="Install reportlab: pip install reportlab"
            )

    st.markdown("---")
    st.markdown(f"**Preview** — First 5 of {len(filtered_df)} filtered leads:")
    preview_cols = ["name", "phone", "email", "category", "validation_status"]
    available_preview = [c for c in preview_cols if c in filtered_df.columns]
    st.dataframe(filtered_df[available_preview].head(5), hide_index=True)
